from flask import Flask, request, render_template, send_file
import openpyxl
import io

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
	# Check if the password is correct
	password = request.form['password']
	if password != '123456':
		return "Invalid password"

	# Check if a file was uploaded
	if 'file' not in request.files:
		return "No file uploaded"

	file = request.files['file']

	# Load the uploaded Excel file
	workbook = openpyxl.load_workbook(file)

	# Apply the code to the workbook

	# Select the desired sheet
	sheet = workbook['Sheet1']

	# Specify the column name of the existing column
	existing_column_name = 'EXPIRY_DATE'

	# Determine the index of the existing column
	existing_column_index = None
	for cell in sheet[1]:
		if cell.value == existing_column_name:
			existing_column_index = cell.column
			break

	# Specify the column name for the new column
	new_column_name = 'Darmound'
	new_column_name2 = 'Expired/Active'

	# Insert the new column next to the existing column in one line
	sheet.insert_cols(existing_column_index + 1)
	sheet.insert_cols(existing_column_index + 2)


	# Rename the newly inserted column
	sheet.cell(row=1, column=existing_column_index + 1, value=new_column_name)
	sheet.cell(row=1, column=existing_column_index + 2, value=new_column_name2)

	new_column_index = None
	for cell in sheet[1]:
		if cell.value == new_column_name:
			new_column_index = cell.column
			break

	new_column_index2 = None
	for cell in sheet[1]:
		if cell.value == new_column_name2:
			new_column_index2 = cell.column
			break
	# Specify the column names
	column_A_name = openpyxl.utils.get_column_letter(existing_column_index)
	column_B_name = openpyxl.utils.get_column_letter(new_column_index)
	column_C_name = openpyxl.utils.get_column_letter(new_column_index2)
	# Get the columns by their names
	column_A = sheet[column_A_name]
	column_B = sheet[column_B_name]
	column_C = sheet[column_C_name]

	# Iterate over the rows in column A, starting from row 2
	for cell_A, cell_B, cell_C in zip(column_A[1:], column_B[1:], column_C[1:]):
		cell_B.value = f'=IF(TODAY() - {cell_A.coordinate} < 0, IF(TODAY() - {cell_A.coordinate} < -90, "valid", IF(TODAY() - {cell_A.coordinate} > -90, "expired soon", "valid")), IF(TODAY() - {cell_A.coordinate} > 0, IF(TODAY() - {cell_A.coordinate} >= 1825, "darmound", "expired"), "expired"))'
		cell_C.value = f'=IF(TODAY() - {cell_A.coordinate} < 0, "valid", "expired")'

	# Save the modified workbook
	workbook.save('updated-data.xlsx')

	# Save the modified workbook to a memory buffer
	buffer = io.BytesIO()
	workbook.save(buffer)
	buffer.seek(0)

	# Send the modified file as a response for download
	return send_file(
		buffer,
		attachment_filename='updated-data.xlsx',
		as_attachment=True
	)


	return "File uploaded and processed successfully"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)

